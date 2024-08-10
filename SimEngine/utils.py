from dataclasses import dataclass
from typing import Dict, List, Optional
from scipy.sparse import csr_matrix

import numpy as np
import pandas as pd


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pprint import pprint


@dataclass
class SimilarityDict:
    x1: List[str]
    x2: List[str]
    similarity_matrix: csr_matrix
    similarity_dict: Dict[str, Dict[str, float]] = None
    threshold: Optional[float] = None
    top_k: Optional[int] = None
    
    def __post_init__(self):
        self._construct_similarity_dict()

    def _construct_similarity_dict(self):
        self.similarity_dict = {}
        for i, xi in enumerate(self.x1):
            item_similarity = [(self.x2[j], self.similarity_matrix[i, j]) for j in range(len(self.x2)) if self.similarity_matrix[i, j] > 0]
            item_similarity.sort(key=lambda x: x[1], reverse=True)
            
            if self.top_k:
                item_similarity = item_similarity[:self.top_k]
            
            if self.threshold:
                item_similarity = [(key, val) for key, val in item_similarity if val >= self.threshold]
            
            self.similarity_dict[xi] = dict(item_similarity)

    def __getitem__(self, key: str) -> Dict[str, float]:
        return self.similarity_dict[key]
    
class BatchGenerator:
    def __init__(self, data: List[str], batch_size: int = 32):
        self.data = data
        self.batch_size = batch_size

    def __iter__(self):
        for i in range(0, len(self.data), self.batch_size):
            yield self.data[i:i + self.batch_size]
            
            
   
   # -- convert similarity dict to excel -- #
           
def get_max_3_scores_min_3_scores_and_2_from_middle(contract_dict: dict[str, float]):
    scores = list(contract_dict.values())
    scores_sorted = sorted(scores)
    
    max_similar_contracts = {k: v for k, v in contract_dict.items() if v in scores_sorted[-5:]}
    min_similar_contracts = {k: v for k, v in contract_dict.items() if v in scores_sorted[:5]}
    middle_scores = scores_sorted[3:-3]
    middle_contracts = {k: v for k, v in contract_dict.items() if v in middle_scores[:3]}
    
    return {
        'max_similar_contracts': max_similar_contracts,
        'min_similar_contracts': min_similar_contracts,
        'middle_contracts': middle_contracts
    }

def convert_sim_dict_to_excel(sim_dict: SimilarityDict, save_dir: str):

    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    new_similarity_dict = {key: get_max_3_scores_min_3_scores_and_2_from_middle(val) for key, val in sim_dict.similarity_dict.items()}

    # Define fill colors for min, middle, and max similarities
    min_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    middle_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    max_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

    # Iterate through the data and write to the Excel sheet with cell coloring
    row_idx = 1
    for fc, similarities in new_similarity_dict.items():
        row_idx += 1
        for category, contracts in similarities.items():
            for sub_contract, score in contracts.items():
                ws.cell(row=row_idx, column=1, value=fc)
                ws.cell(row=row_idx, column=2, value=sub_contract)
                ws.cell(row=row_idx, column=3, value=score)
                
                if category == 'min_similar_contracts':
                    ws.cell(row=row_idx, column=3).fill = min_fill
                elif category == 'middle_contracts':
                    ws.cell(row=row_idx, column=3).fill = middle_fill
                elif category == 'max_similar_contracts':
                    ws.cell(row=row_idx, column=3).fill = max_fill
                
                row_idx += 1

    wb.save(save_dir)


if __name__ == '__main__':
    pkled = pd.read_pickle('data/similarity_dict_1.pkl')
    print(pkled)
    convert_sim_dict_to_excel(pkled, 'data/similarity_dict_1.xlsx')